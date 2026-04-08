import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


def export_to_excel(results, filepath=None):
    if filepath is None:
        reports_dir = os.path.join(os.path.dirname(
            os.path.dirname(__file__)), 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(
            reports_dir, f"gas_forecast_report_{timestamp}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    # ==================== ЛИСТ 1: Входные параметры ====================
    ws_params = wb.create_sheet("Входные параметры")
    ws_params['A1'] = "ПАРАМЕТРЫ РАСЧЕТА"
    ws_params['A1'].font = Font(size=14, bold=True)
    ws_params.merge_cells('A1:B1')

    params = results['params']
    param_list = [
        ("Начальное пластовое давление, МПа", params.get('P_pl', 0)),
        ("Пластовая температура, °C", params.get('T_pl', 0)),
        ("Начальные запасы газа, млрд м³", params.get('G_nach', 0)),
        ("Относительная плотность газа", params.get('rho_otn', 0)),
        ("Количество скважин, шт", params.get('N_skv', 0)),
        ("Глубина скважины, м", params.get('H_skv', 0)),
        ("Диаметр НКТ, мм", params.get('d_NKT', 0)),
        ("Коэффициент a", params.get('a_coef', 0)),
        ("Коэффициент b", params.get('b_coef', 0)),
        ("Макс. депрессия, МПа", params.get('dP_max', 0)),
        ("Коэффициент эксплуатации", params.get('K_eks', 0)),
        ("Режим ДКС", params.get('DKS_mode', '')),
        ("Давление на входе ДКС, МПа", params.get('P_vh_DKS', 0)),
        ("Давление на выходе ДКС, МПа", params.get('P_vyh_DKS', 0)),
        ("Мощность ДКС, МВт", params.get('N_DKS', 0)),
        ("Год начала разработки", params.get('start_year', 0)),
        ("Горизонт прогноза, лет", params.get('T_max', 0)),
        ("Уровень полки, млрд м³/год", params.get('Q_polka', 0)),
        ("Начальный ВГФ, г/м³", params.get('VGF_nach', 0)),
        ("Интенсивность роста ВГФ", params.get('dVGF_dG', 0)),
        ("Критический ВГФ, г/м³", params.get('VGF_krit', 0)),
        ("Учет трения", "Да" if params.get('opt_friction', 0) == 1 else "Нет"),
        ("Метод PVT", params.get('pvt_method', ''))
    ]

    for i, (name, value) in enumerate(param_list, start=3):
        ws_params.cell(row=i, column=1, value=name).font = Font(bold=True)
        ws_params.cell(row=i, column=2, value=value)

    # ==================== ЛИСТ 2: Помесячные результаты ====================
    ws_monthly = wb.create_sheet("Помесячные результаты")
    monthly_df = pd.DataFrame(results['monthly_df'])

    for col_idx, col_name in enumerate(monthly_df.columns, 1):
        cell = ws_monthly.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD",
                                end_color="DDDDDD", fill_type="solid")

    for row_idx, row in enumerate(monthly_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws_monthly.cell(row=row_idx, column=col_idx, value=value)

    # ==================== ЛИСТ 3: Годовые результаты ====================
    ws_yearly = wb.create_sheet("Годовые результаты")
    yearly_df = pd.DataFrame(results['yearly_df'])

    for col_idx, col_name in enumerate(yearly_df.columns, 1):
        cell = ws_yearly.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD",
                                end_color="DDDDDD", fill_type="solid")

    for row_idx, row in enumerate(yearly_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws_yearly.cell(row=row_idx, column=col_idx, value=value)

    # ==================== АВТОШИРИНА КОЛОНОК (ИСПРАВЛЕНО) ====================
    for ws in [ws_params, ws_monthly, ws_yearly]:
        for col in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(filepath)
    return filepath
